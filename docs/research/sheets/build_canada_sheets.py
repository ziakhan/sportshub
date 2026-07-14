#!/usr/bin/env python3
"""Build canada-basketball-research-2026-07.xlsx from the Canada census data
(canada-census-data.json, saved from the 2026-07-14 deep-research run).
Sheets: Overview, Governing Bodies, Leagues, All Clubs, Clubs <prov> x10, Verification.
No network access — reorganizes existing data only."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = os.path.dirname(os.path.abspath(__file__))
data = json.load(open(os.path.join(D, "canada-census-data.json")))["result"]["provinces"]

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

def sheet(wb, title, headers, rows, widths, wrap_cols=()):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
    for r in rows:
        ws.append([("" if v is None else v) for v in r])
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP if cell.column in wrap_cols else TOP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws

def club_row(p, c):
    return [c.get("name"), p["province"], c.get("region"), c.get("city"), c.get("address"),
            c.get("phone"), c.get("email"), c.get("website"), c.get("principal"),
            c.get("programs"), "; ".join(c.get("leagues") or []), c.get("team_league_map"),
            c.get("type"), c.get("software"), c.get("confidence"),
            (c.get("source_urls") or [""])[0], c.get("notes")]

CLUB_HDR = ["Club Name", "Province", "Region", "City", "Address", "Phone", "Email", "Website",
            "Owner/Principal", "Programs", "Leagues", "Team-League Map", "Type", "Software",
            "Confidence", "Source", "Notes"]
CLUB_W = [34, 12, 20, 18, 28, 16, 26, 30, 26, 32, 26, 30, 18, 26, 11, 34, 40]
CLUB_WRAP = (10, 12, 17)

wb = Workbook()
wb.remove(wb.active)

# ---- Overview ----
PLATFORMS = ["RAMP", "TeamLinkt", "TeamSnap", "SportsEngine", "Spordle", "Kreezee", "Exposure",
             "LeagueApps", "eSportsDesk", "itSportsNet", "GOALLINE", "LeagueToolbox", "Citrus",
             "bracketteam", "Amilia", "Zeffy", "EventNroll", "Wix", "WordPress", "Squarespace"]
ov_rows = []
for p in data:
    clubs, leagues, verifs = p["clubs"], p["leagues"], p.get("verification", [])
    ov_rows.append([p["province"], len(leagues), len(clubs),
                    sum(1 for c in clubs if c.get("phone") or c.get("email")),
                    sum(1 for c in clubs if c.get("principal")),
                    sum(1 for v in verifs if v["verdict"] == "confirmed"),
                    sum(1 for v in verifs if v["verdict"] == "refuted"),
                    "" if verifs else "no verification pass (session limits)"])
tot = ["TOTAL", sum(r[1] for r in ov_rows), sum(r[2] for r in ov_rows), sum(r[3] for r in ov_rows),
       sum(r[4] for r in ov_rows), sum(r[5] for r in ov_rows), sum(r[6] for r in ov_rows), ""]
ws = sheet(wb, "Overview",
           ["Province", "Leagues", "Clubs", "Clubs w. direct contact", "Clubs w. named principal",
            "Spot-checks confirmed", "Refuted (corrected)", "Coverage note"],
           ov_rows + [tot], [26, 10, 9, 20, 21, 19, 17, 38])
ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

# software landscape block below the table
start = ws.max_row + 2
ws.cell(row=start, column=1, value="Software landscape (mentions)").font = Font(bold=True)
counts = {}
for p in data:
    for kind, items in (("league", p["leagues"]), ("club", p["clubs"])):
        for x in items:
            sw = (x.get("software") or "").lower()
            for pl in PLATFORMS:
                if pl.lower() in sw:
                    counts.setdefault(pl, [0, 0])[0 if kind == "league" else 1] += 1
ws.cell(row=start + 1, column=1, value="Platform").font = Font(bold=True)
ws.cell(row=start + 1, column=2, value="Leagues").font = Font(bold=True)
ws.cell(row=start + 1, column=3, value="Clubs").font = Font(bold=True)
for i, (pl, (lg, cl)) in enumerate(sorted(counts.items(), key=lambda kv: -(kv[1][0] + kv[1][1]))):
    ws.cell(row=start + 2 + i, column=1, value=pl)
    ws.cell(row=start + 2 + i, column=2, value=lg)
    ws.cell(row=start + 2 + i, column=3, value=cl)

# ---- Governing Bodies ----
sheet(wb, "Governing Bodies",
      ["Province", "PSO", "Website", "Phone", "Email", "Address", "Member registry", "Software", "Notes"],
      [[p["province"], p["pso"].get("name"), p["pso"].get("website"), p["pso"].get("phone"),
        p["pso"].get("email"), p["pso"].get("address"), p["pso"].get("registry_url"),
        p["pso"].get("software"), p["pso"].get("notes")] for p in data],
      [24, 34, 30, 16, 26, 30, 40, 40, 60], wrap_cols=(8, 9))

# ---- Leagues ----
sheet(wb, "Leagues",
      ["Province", "League", "Operator", "Level", "Ages/Tiers", "Region", "Scale", "Season",
       "Software", "Contact", "Website", "Confidence", "Notes"],
      [[p["province"], l.get("name"), l.get("operator"), l.get("level"), l.get("ages_tiers"),
        l.get("region"), l.get("scale"), l.get("season"), l.get("software"), l.get("contact"),
        l.get("website"), l.get("confidence"), l.get("notes")]
       for p in data for l in p["leagues"]],
      [14, 34, 30, 15, 26, 20, 34, 24, 34, 26, 30, 11, 50], wrap_cols=(3, 5, 7, 9, 13))

# ---- All Clubs + per-province ----
sheet(wb, "All Clubs", CLUB_HDR, [club_row(p, c) for p in data for c in p["clubs"]],
      CLUB_W, CLUB_WRAP)
SHORT = {"Quebec": "QC", "British Columbia": "BC", "Alberta": "AB", "Manitoba": "MB",
         "Saskatchewan": "SK", "Nova Scotia": "NS", "New Brunswick": "NB",
         "Newfoundland & Labrador": "NL", "Prince Edward Island": "PEI",
         "Territories (YT/NWT/NU)": "Territories"}
for p in data:
    rows = sorted((club_row(p, c) for c in p["clubs"]), key=lambda r: (r[3] or "", r[0] or ""))
    sheet(wb, f"Clubs {SHORT[p['province']]}", CLUB_HDR, rows, CLUB_W, CLUB_WRAP)

# ---- Verification ----
sheet(wb, "Verification",
      ["Province", "Verdict", "Claim", "Evidence", "Correction"],
      [[p["province"], v.get("verdict", "").upper(), v.get("claim"), v.get("evidence"),
        v.get("correction")] for p in data for v in p.get("verification", [])],
      [22, 12, 60, 70, 50], wrap_cols=(3, 4, 5))

out = os.path.join(D, "canada-basketball-research-2026-07.xlsx")
wb.save(out)
print(out, "-", len(wb.sheetnames), "sheets:", ", ".join(wb.sheetnames))
